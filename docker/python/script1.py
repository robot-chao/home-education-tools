# 创建一个Excel，生成20到计算题
from openpyxl import Workbook
import random
from uitls import genAllArithmeticItems, write2Excel

# 生成的题目中包含0的题目太多了

# 30道20以内加减法
paperBook = Workbook()
paperSheet = paperBook.active
paperSheet.title = '四则运算'

itemCount = 30

allItems = genAllArithmeticItems(20, 1, (1,2))
for i in range(10):
    random.shuffle(allItems)
    nextRow = write2Excel(paperSheet, allItems[:30], 3)

paperBook.save('20以内加减法计算题.xlsx')

# 30道100以内加减法
paperBook = Workbook()
paperSheet = paperBook.active
paperSheet.title = '四则运算'

itemCount = 30

allItems = genAllArithmeticItems(100, 20)
for i in range(10):
    random.shuffle(allItems)
    nextRow = write2Excel(paperSheet, allItems[:30], 3)

paperBook.save('100以内加减法计算题.xlsx')

# 30道10以内加减法
paperBook = Workbook()
paperSheet = paperBook.active
paperSheet.title = '四则运算'

itemCount = 30
allItems = genAllArithmeticItems(10, 1, (1,2))
for i in range(6):
    random.shuffle(allItems)
    write2Excel(paperSheet, allItems[:30], 3)
paperSheet.page_margins.bottom=0.8
paperBook.save('10以内加减法计算题.xlsx')

# 30道10以内加减法，填加数，减数
paperBook = Workbook()
paperSheet = paperBook.active
paperSheet.title = '四则运算'
# print(paperSheet.max_row)
itemCount = 30
allItems = genAllArithmeticItems(10, 1)
for i in range(6):
    random.shuffle(allItems)
    write2Excel(paperSheet, allItems[:30], 3)
    # print(paperSheet.max_row)

paperSheet.page_margins.bottom=0.8
paperBook.save('10以内加减法计算题V2.xlsx')