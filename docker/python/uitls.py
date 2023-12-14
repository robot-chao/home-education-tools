from openpyxl.styles import *
from openpyxl.utils import get_column_letter
import random
import math

# 生成题目
def genArithmeticItem(type, max=10, min=0):
    # 取一个min~max之间的数作为和
    sum = random.randint(min, max)
    # 在min和sum之前取一个数作为加数
    addend = random.randint(min, (sum - min) if sum - min > min else sum)
    if 1 == type:
        return "{} + {} = ".format(addend, sum - addend)
    else:
        return "{} - {} = ".format(sum, addend)

def genAllArithmeticItems(max=10, min=0, types=(1,2,3,4,5,6)):
    allItems = []
    for sum in range(min, max + 1):
        for addend in range(min, sum):
            if 1 in types:
                allItems.append("{} + {} = ".format(addend, sum - addend))
            if 2 in types:
                allItems.append("{} - {} = ".format(sum, addend))
            if 3 in types:
                allItems.append("(    ) + {} = {}".format(sum - addend, sum))
            if 4 in types:
                allItems.append("{} + (    ) = {}".format(addend, sum))
            if 5 in types:
                allItems.append("(    ) - {} = {}".format(addend, sum - addend))
            if 6 in types:
                allItems.append("{} - (    ) = {}".format(sum, sum - addend))
    return allItems

# 生成试卷
def genArithmeticPaper(paperSheet, itemCount=30, startRow=1, columnNumPerRow=3, maxNum=20, minNum=0, types=(1,2,3,4,5,6)):
    # 生成所有的题目
    allItems = genAllArithmeticItems(maxNum, minNum, types)

    # 随机打乱顺序
    random.shuffle(allItems)
    # 取前面itemCont个
    items = allItems[:itemCount]

    font0 = Font(name='微软雅黑', size=14, vertAlign='baseline')
    curRow = startRow
    rowCount = math.ceil(itemCount / columnNumPerRow)
    for i in range(1, rowCount + 1):
        for i in range(1, columnNumPerRow + 1):
            column = i % columnNumPerRow + 1
            paperSheet.cell(curRow, column, value= items.pop())
            paperSheet.cell(curRow, column).font = font0
            if column == 1:
                paperSheet.row_dimensions[curRow].height = 30
            if curRow == 1:
                paperSheet.column_dimensions[get_column_letter(column)].width = 86 / columnNumPerRow
        curRow += 1

    paperSheet.merge_cells(start_row=curRow, start_column=1, end_row=curRow, end_column=columnNumPerRow)
    paperSheet.cell(curRow, 1).value='日期：_________，用时：________，得分：________'
    paperSheet.cell(curRow, 1).alignment = Alignment(horizontal='right',vertical='center')
    paperSheet.row_dimensions[curRow].height = 50

    # 返回下一行号
    return curRow + 1

# 生成试卷
def write2Excel(paperSheet, items=[], columnNumPerRow=3):
    itemCount = len(items)

    font0 = Font(name='微软雅黑', size=14, vertAlign='baseline')
    curRow = paperSheet.max_row

    if curRow > 1:
        curRow = curRow + 1
    # print(curRow)
    rowCount = math.ceil(itemCount / columnNumPerRow)
    for i in range(1, rowCount + 1):
        for i in range(1, columnNumPerRow + 1):
            column = i % columnNumPerRow + 1
            paperSheet.cell(curRow, column, value= items.pop())
            paperSheet.cell(curRow, column).font = font0
            if column == 1:
                paperSheet.row_dimensions[curRow].height = 30
            if curRow == 1:
                paperSheet.column_dimensions[get_column_letter(column)].width = 86 / columnNumPerRow
        curRow += 1

    paperSheet.merge_cells(start_row=curRow, start_column=1, end_row=curRow, end_column=columnNumPerRow)
    paperSheet.cell(curRow, 1).value='日期：_________，用时：________，得分：________'
    paperSheet.cell(curRow, 1).alignment = Alignment(horizontal='right',vertical='center')
    paperSheet.row_dimensions[curRow].height = 50

    # 返回下一行号
    return curRow + 1

